from openpyxl import load_workbook

excel_sheet = load_workbook(filename = "TestExcelSheet.xlsx")
SECTION_INDEX = 1
LESSON_INDEX = 2
SOURCE_INDEX = 3

same_source_lesson = []
modified_source_lesson = []

def main():
    excel_sheet.active = 0
    # sheets = excel_sheet.active
    # print(sheets.title)
    # to print a specific cell: 
    # print(sheet['A:C'], values_only=True)

    #USER INPUT GETTERS:
    # course_name = input("Which course are you working with? ")
    # lesson_name = input(f"What lesson are you looking for in {course_name}? ") 
    lesson_name = 'MyPlate'
    sheet_name = 'Exercise Science Duplicate'
    user_source = retrieve_user_section(lesson_name, sheet_name)
    for sheet in excel_sheet.worksheets: #iterates through the sheets
        print(sheet.title) #prints the name of the title
        '''I'm 90% sure I won't need this, I just want to be sure before I delete it all lol'''
        # for row in sheet.iter_rows(min_col=LESSON_INDEX, max_col=LESSON_INDEX): #iterates through the rows of the sheet
        #     cell = row[0] #allows the cell to be read properly
        #     if cell.value != None:
        #         if cell.value.lower() == lesson_name.lower():
        #             print(f"{cell.value} fount at: \nRow: {cell.row}")
        #             print(retrieving_section(cell.row, sheet))
        retrieved_new_section = iterate_through_rows(sheet, lesson_name, user_source)

            
'''
In terms of functions:
    - retrieve section number
    - row iterator
    - check against word
    - if word == user input
    - check SOURCE index if it matches add to list
        - is modified
        - is original
'''
def iterate_through_rows(sheet, lesson_name, original_source):
    for row in sheet.iter_rows(min_col=LESSON_INDEX, max_col=LESSON_INDEX): #iterates through the rows of the sheet
        cell = row[0] #allows the cell to be read properly
        if cell.value != None:
            if cell.value.lower() == lesson_name.lower():
                # print(f"{cell.value} fount at: \nRow: {cell.row}")
                source_bool, retrieved_section = checking_source(cell.row, sheet, original_source)
                if source_bool == True:
                    #so actually I think that this needs to change so it can be used in retrieve user section
                    same_source_lesson.append(retrieved_section)
                elif source_bool == False:
                    modified_source_lesson.append(retrieved_section)
                #return retrieving_section(cell.row, sheet)#returns the source of the derived lesson
    return None


def retrieve_user_section(lesson_name, sheet_name):
    sheet = excel_sheet[sheet_name]
    user_section = iterate_through_rows(sheet, lesson_name)
    sheet_code = sheet['A3'].value
    user_source = f"{sheet_code} {split_section(user_section)}"
    return user_source

def retrieving_section(row_number, sheet):
    #Getting the source of derived course
    row=row_number
    cell = sheet.cell(row=row, column=SECTION_INDEX)
    while cell.value == None:
        if cell.value == None:
            row = row - 1
        else:
            break   
        cell = sheet.cell(row=row, column=SECTION_INDEX)
        #need to update this so it returns the course & the section name
    sheet_code = sheet['A3'].value
    source_code = f"{sheet_code} {split_section(cell.value)}"
    return source_code

def checking_source(source_row, sheet, original_source):
    cell = sheet.cell(row = source_row, column = SOURCE_INDEX)
    cell_contents = cell.value
    if is_original(cell_contents):
        return None, None
    # elif is_modified(cell_contents):
    #     cell_pieces = cell_contents.split() 
    #     to_add_to_modified = f"{cell_pieces[0]} {cell_pieces[1]}"
    #     modified_source_lesson.append(to_add_to_modified)
    #     return False
    elif cell_contents == original_source:
        return True, retrieving_section(cell.row, sheet)
    elif is_modified(cell_contents):
        return False, retrieving_section(cell.row, sheet)
    else:
        return None, None

def is_modified(cell_contents):
    cell = cell_contents.strip().lower()
    split_cell = cell.split()
    if "modified" in split_cell:
        return True
    else: 
        return False

def is_original(cell_contents):
    cell = cell_contents.strip().lower()
    split_cell = cell.split()
    if "original" in split_cell:
        return True
    else: 
        return False
    
def split_section(section):
    #will have to be able to split the section code by spaces and colons
    cleaned = section.replace(":", "")
    split_section = cleaned.split()
    return split_section[1]


if __name__ == "__main__":
    main()

# to access the title of the sheet you can set the original sheet = excel_sheet.active then iterate through sheets and make "sheet.title" the dictionaries being stored
   

# column_values_list = []
# for cell in cell.column[3]:
#     print(cell.value, end=" ")

# def iter_excel_openpyxl(file: IO[bytes]) -> Iterator[dict[str, object]]:
#     workbook = openpyxl.load_workbook(file, read_only=True)
#     rows = workbook.active.rows
#     headers = [str(cell.value) for cell in next(rows)]
#     for row in rows:
#         yield dict(zip(headers, (cell.value for cell in row)))

# with open(excel_sheet, 'rb') as f:
#     rows = iter_excel_openpyxl(f)
#     row = next(rows)
#     print(row)
    

# if __name__ == "__iter_excel_openpyxl__":
#     inter_excel_openpyxl()